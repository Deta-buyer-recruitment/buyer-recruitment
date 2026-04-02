from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional
import io
from datetime import datetime

from services.supabase_client import get_supabase

router = APIRouter()


@router.get("/")
async def list_buyers(customer_id: Optional[str] = None, status: Optional[str] = None):
    sb = get_supabase()
    q = sb.table("buyers").select("*, customers(name), contact_logs(*)").order("no")
    if customer_id:
        q = q.eq("customer_id", customer_id)
    if status:
        q = q.eq("status", status)
    return q.execute().data


@router.get("/{buyer_id}")
async def get_buyer(buyer_id: str):
    sb = get_supabase()
    result = sb.table("buyers").select("*, customers(name), contact_logs(*)").eq("id", buyer_id).single().execute()
    if not result.data:
        raise HTTPException(status_code=404, detail="바이어 없음")
    return result.data


class BuyerUpdate(BaseModel):
    company: Optional[str] = None
    country: Optional[str] = None
    website: Optional[str] = None
    contact_name: Optional[str] = None
    position: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    contact_name2: Optional[str] = None
    position2: Optional[str] = None
    email2: Optional[str] = None
    contact_name3: Optional[str] = None
    position3: Optional[str] = None
    email3: Optional[str] = None
    summary: Optional[str] = None
    notes: Optional[str] = None
    follow_up: Optional[str] = None
    status: Optional[str] = None


@router.patch("/{buyer_id}")
async def update_buyer(buyer_id: str, payload: BuyerUpdate):
    sb = get_supabase()
    data = {k: v for k, v in payload.dict().items() if v is not None}
    if not data:
        raise HTTPException(status_code=400, detail="수정할 내용 없음")
    result = sb.table("buyers").update(data).eq("id", buyer_id).execute()
    return result.data[0]


@router.get("/export/excel")
async def export_excel(customer_id: str):
    """
    기존 Contact Log 엑셀 양식과 완전히 동일한 형태로 Export
    컬럼: 고객사명, No., 바이어사명, 국가, 웹사이트, 담당자명, 직함, 이메일, 전화번호,
          주요내용, 비고, 1차~10차 연락(날짜/방법/회신여부/결과), Follow-up
    contact2/3도 별도 행으로 포함
    """
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sb = get_supabase()
    buyers = sb.table("buyers").select(
        "*, customers(name), contact_logs(*)"
    ).eq("customer_id", customer_id).order("no").execute().data or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buyer_List & LOG"

    # ── 헤더 (기존 양식과 동일) ──
    headers = [
        "고객사명", "No.", "바이어사명 (Company)", "국가 (Country)", "웹사이트 (Website)",
        "담당자명 (Contact Name)", "직함 (Position)", "이메일", "전화번호 (Phone)",
        "주요 내용 (Summary)", "비고 (Notes)",
    ]
    for i in range(1, 11):
        suffix = "" if i == 1 else f".{i-1}"
        date_col = f"{i}차 연락 일자 ( Date)"
        method_col = f"연락방법{suffix}"
        reply_col  = f"회신/응답 여부 (Y/N){suffix}"
        result_col = f"연락 결과{suffix}"
        headers += [date_col, method_col, reply_col, result_col]
    headers += ["주요 내용 (Summary).1", "Follow-up", "비고 (Notes).1"]

    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 30

    # ── 데이터 행 ──
    row_idx = 2
    for buyer in buyers:
        logs = {log["attempt_no"]: log for log in (buyer.get("contact_logs") or [])}
        customer_name = (buyer.get("customers") or {}).get("name", "")

        # contact1 (메인)
        contacts = [
            (buyer.get("contact_name"), buyer.get("position"), buyer.get("email"), buyer.get("phone")),
        ]
        # contact2, contact3 도 있으면 추가
        if buyer.get("email2"):
            contacts.append((buyer.get("contact_name2"), buyer.get("position2"), buyer.get("email2"), None))
        if buyer.get("email3"):
            contacts.append((buyer.get("contact_name3"), buyer.get("position3"), buyer.get("email3"), None))

        for ci, (cname, cpos, cemail, cphone) in enumerate(contacts):
            row_data = [
                customer_name,
                buyer.get("no", ""),
                buyer.get("company", ""),
                buyer.get("country", ""),
                buyer.get("website", ""),
                cname or "",
                cpos or "",
                cemail or "",
                cphone or buyer.get("phone", "") if ci == 0 else "",
                buyer.get("summary", "") if ci == 0 else "",
                buyer.get("notes", "")   if ci == 0 else "",
            ]

            # 1차~10차 연락 (contact1에만 기록)
            for attempt in range(1, 11):
                log = logs.get(attempt, {}) if ci == 0 else {}
                row_data += [
                    log.get("contact_date", ""),
                    log.get("contact_method", ""),
                    "Y" if log.get("replied") is True else ("N" if log.get("replied") is False else ""),
                    log.get("result", ""),
                ]

            row_data += [
                buyer.get("summary", "") if ci == 0 else "",
                buyer.get("follow_up", "") if ci == 0 else "",
                buyer.get("notes", "")   if ci == 0 else "",
            ]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            row_idx += 1

    # 열 너비 조정
    col_widths = {
        "A": 10, "B": 6,  "C": 28, "D": 12, "E": 30,
        "F": 18, "G": 18, "H": 28, "I": 18, "J": 30, "K": 20,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    # 연락 컬럼들 너비
    for i in range(12, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 14

    # 행 높이
    for r in range(2, row_idx):
        ws.row_dimensions[r].height = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Buyer_List_Contact_Log_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.delete("/{buyer_id}")
async def delete_buyer(buyer_id: str):
    sb = get_supabase()
    # 컨택 로그 먼저 삭제
    sb.table("contact_logs").delete().eq("buyer_id", buyer_id).execute()
    # 바이어 삭제
    sb.table("buyers").delete().eq("id", buyer_id).execute()
    return {"success": True}
