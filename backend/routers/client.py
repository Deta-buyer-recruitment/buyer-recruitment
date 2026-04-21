"""
고객용 API 라우터
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from typing import Optional
import os
import hashlib
from datetime import datetime

from services.supabase_client import get_supabase

router = APIRouter()
STORAGE_BUCKET = "project-files"


# ── 고객 로그인 (아이디/비번 검증) ──────────────────────────

class ClientLoginPayload(BaseModel):
    slug: str
    access_id: str
    access_password: str

@router.post("/login")
async def client_login(payload: ClientLoginPayload):
    """고객 대시보드 로그인 — slug + 아이디/비번 검증"""
    sb = get_supabase()
    result = sb.table("customers").select(
        "id, name, slug, access_id, access_password"
    ).eq("slug", payload.slug).execute()

    if not result.data or len(result.data) == 0:
        raise HTTPException(status_code=404, detail="Client not found")

    c = result.data[0]

    # 아이디/비번 미설정 시 접근 차단
    if not c.get("access_id") or not c.get("access_password"):
        raise HTTPException(status_code=403, detail="Access not configured. Contact your project manager.")

    if c["access_id"] != payload.access_id:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    # 비번 비교: Python hashlib.sha256 == pgcrypto encode(digest(),'hex')
    stored_pw    = c["access_password"]
    input_pw     = payload.access_password
    hashed_input = hashlib.sha256(input_pw.encode()).hexdigest()

    if stored_pw != input_pw and stored_pw != hashed_input:
        raise HTTPException(status_code=401, detail="Invalid credentials")

    return {"success": True, "customer_id": c["id"], "customer_name": c["name"]}


# ── 타임라인 ─────────────────────────────────────────────────

@router.get("/timeline/{customer_id}")
async def get_timeline(customer_id: str):
    sb = get_supabase()
    result = sb.table("timelines").select("*").eq(
        "customer_id", customer_id
    ).order("step_no").execute()
    if not result.data:
        sb.rpc("init_timeline", {"p_customer_id": customer_id}).execute()
        result = sb.table("timelines").select("*").eq(
            "customer_id", customer_id
        ).order("step_no").execute()
    return result.data


class TimelineUpdate(BaseModel):
    step_no: int
    status: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    memo: Optional[str] = None


@router.patch("/timeline/{customer_id}")
async def update_timeline(customer_id: str, payload: TimelineUpdate):
    sb = get_supabase()
    data = {k: v for k, v in payload.dict().items() if v is not None and k != "step_no"}
    data["updated_at"] = datetime.now().isoformat()
    result = sb.table("timelines").update(data).eq(
        "customer_id", customer_id
    ).eq("step_no", payload.step_no).execute()
    return result.data[0]


class TimelineStepAdd(BaseModel):
    step_name: str
    start_date: Optional[str] = None
    end_date: Optional[str] = None


@router.post("/timeline/{customer_id}/add-step")
async def add_timeline_step(customer_id: str, payload: TimelineStepAdd):
    sb = get_supabase()
    existing = sb.table("timelines").select("step_no").eq(
        "customer_id", customer_id
    ).order("step_no", desc=True).limit(1).execute().data
    next_step = (existing[0]["step_no"] + 1) if existing else 7
    result = sb.table("timelines").insert({
        "customer_id": customer_id,
        "step_no":     next_step,
        "step_name":   payload.step_name,
        "start_date":  payload.start_date,
        "end_date":    payload.end_date,
        "status":      "pending",
    }).execute()
    return result.data[0]


@router.delete("/timeline/{customer_id}/step/{step_no}")
async def delete_timeline_step(customer_id: str, step_no: int):
    if step_no <= 6:
        raise HTTPException(status_code=400, detail="Default 6 steps cannot be deleted")
    sb = get_supabase()
    sb.table("timelines").delete().eq("customer_id", customer_id).eq("step_no", step_no).execute()
    return {"deleted": True}


# ── 파일 ─────────────────────────────────────────────────────

@router.get("/files/{customer_id}")
async def list_files(customer_id: str):
    sb = get_supabase()
    result = sb.table("project_files").select("*").eq(
        "customer_id", customer_id
    ).order("created_at", desc=True).execute()
    return result.data


@router.post("/files/{customer_id}/upload")
async def upload_file(
    customer_id: str,
    category: str = Form("report"),
    uploader_id: str = Form(...),
    file: UploadFile = File(...)
):
    sb = get_supabase()
    contents = await file.read()
    storage_path = f"{customer_id}/{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    sb.storage.from_(STORAGE_BUCKET).upload(
        storage_path, contents,
        {"content-type": file.content_type or "application/octet-stream"}
    )
    result = sb.table("project_files").insert({
        "customer_id": customer_id,
        "name": file.filename,
        "category": category,
        "storage_path": storage_path,
        "size_bytes": len(contents),
        "uploaded_by": uploader_id,
    }).execute()
    return result.data[0]


@router.get("/files/{customer_id}/download/{file_id}")
async def download_file(customer_id: str, file_id: str):
    sb = get_supabase()
    f = sb.table("project_files").select("storage_path, name").eq(
        "id", file_id
    ).eq("customer_id", customer_id).single().execute().data
    if not f:
        raise HTTPException(status_code=404, detail="File not found")
    signed = sb.storage.from_(STORAGE_BUCKET).create_signed_url(f["storage_path"], 3600)
    return {"signed_url": signed["signedURL"], "filename": f["name"]}


@router.delete("/files/{customer_id}/{file_id}")
async def delete_file(customer_id: str, file_id: str):
    sb = get_supabase()
    f = sb.table("project_files").select("storage_path").eq("id", file_id).single().execute().data
    if not f:
        raise HTTPException(status_code=404, detail="File not found")
    sb.storage.from_(STORAGE_BUCKET).remove([f["storage_path"]])
    sb.table("project_files").delete().eq("id", file_id).execute()
    return {"deleted": True}


# ── 미팅 ─────────────────────────────────────────────────────

@router.get("/meetings/{customer_id}")
async def list_meetings(customer_id: str):
    sb = get_supabase()
    result = sb.table("meetings").select("*, buyers(company, country)").eq(
        "customer_id", customer_id
    ).order("meeting_date").execute()
    return result.data


@router.get("/meetings/{customer_id}/count")
async def count_meetings(customer_id: str):
    sb = get_supabase()
    result = sb.table("meetings").select("id", count="exact").eq(
        "customer_id", customer_id
    ).eq("status", "scheduled").execute()
    return {"count": result.count or 0}


class MeetingCreate(BaseModel):
    buyer_id: Optional[str] = None
    title: str
    meeting_date: Optional[str] = None
    location: Optional[str] = None
    notes: Optional[str] = None


@router.post("/meetings/{customer_id}")
async def create_meeting(customer_id: str, payload: MeetingCreate):
    sb = get_supabase()
    result = sb.table("meetings").insert({
        "customer_id": customer_id,
        **payload.dict(exclude_none=True)
    }).execute()
    return result.data[0]


@router.patch("/meetings/{meeting_id}")
async def update_meeting(meeting_id: str, payload: dict):
    sb = get_supabase()
    result = sb.table("meetings").update(payload).eq("id", meeting_id).execute()
    return result.data[0]


# ── 문의 ─────────────────────────────────────────────────────

@router.get("/inquiries/{customer_id}")
async def list_inquiries(customer_id: str):
    sb = get_supabase()
    result = sb.table("inquiries").select("*").eq(
        "customer_id", customer_id
    ).order("created_at", desc=True).execute()
    return result.data


class InquiryCreate(BaseModel):
    author_name: Optional[str] = None
    title: str
    content: str


@router.post("/inquiries/{customer_id}")
async def create_inquiry(customer_id: str, payload: InquiryCreate):
    sb = get_supabase()
    result = sb.table("inquiries").insert({
        "customer_id": customer_id,
        **payload.dict(exclude_none=True)
    }).execute()
    return result.data[0]


class InquiryAnswer(BaseModel):
    answer: str


@router.patch("/inquiries/{inquiry_id}/answer")
async def answer_inquiry(inquiry_id: str, payload: InquiryAnswer):
    sb = get_supabase()
    result = sb.table("inquiries").update({
        "answer": payload.answer,
        "status": "answered",
        "answered_at": datetime.now().isoformat()
    }).eq("id", inquiry_id).execute()
    return result.data[0]


# ── 대시보드 통합 API ─────────────────────────────────────────

async def _build_dashboard(customer_id: str) -> dict:
    from datetime import datetime, timedelta
    sb = get_supabase()
    customer  = sb.table("customers").select("id,name,slug").eq("id", customer_id).single().execute().data
    timeline  = sb.table("timelines").select("*").eq("customer_id", customer_id).order("step_no").execute().data or []
    buyers    = sb.table("buyers").select("id,company,country,status").eq("customer_id", customer_id).execute().data or []
    files     = sb.table("project_files").select("*").eq("customer_id", customer_id).order("created_at", desc=True).execute().data or []
    meetings  = sb.table("meetings").select("*,buyers(company,country)").eq("customer_id", customer_id).order("meeting_date").execute().data or []
    inquiries = sb.table("inquiries").select("*").eq("customer_id", customer_id).order("created_at", desc=True).execute().data or []

    # campaigns 테이블에서 target_country 수집
    campaigns_result = sb.table("campaigns").select("target_country").eq(
        "customer_id", customer_id
    ).execute().data or []
    # 중복 제거, None/빈 값 제외
    target_countries = list(dict.fromkeys(
        c["target_country"] for c in campaigns_result
        if c.get("target_country")
    ))

    # contact_logs 전체 조회 (weekly 집계용)
    buyer_ids = [b["id"] for b in buyers]
    all_logs = []
    if buyer_ids:
        # Supabase in_() URL 길이 제한 → 200개씩 청크로 나눠 조회
        chunk_size = 200
        for i in range(0, len(buyer_ids), chunk_size):
            chunk = buyer_ids[i:i + chunk_size]
            chunk_result = sb.table("contact_logs").select(
                "buyer_id, attempt_no, contact_date, replied"
            ).in_("buyer_id", chunk).limit(10000).execute().data or []
            all_logs.extend(chunk_result)

    if not timeline:
        sb.rpc("init_timeline", {"p_customer_id": customer_id}).execute()
        timeline = sb.table("timelines").select("*").eq("customer_id", customer_id).order("step_no").execute().data or []

    total_steps = max(len(timeline), 1)
    done    = len([t for t in timeline if t["status"] == "done"])
    in_prog = len([t for t in timeline if t["status"] == "in_progress"])
    progress = round((done + in_prog * 0.5) / total_steps * 100)

    # contact_logs 기반 통계
    # 총 컨택 수 = contact_logs 전체 건수 (총 시도 횟수)
    total_contacted = len(all_logs)
    # 회신 수 = replied == True 인 고유 바이어 수
    replied_ids   = set(log["buyer_id"] for log in all_logs if log.get("replied") is True)
    total_replied   = len(replied_ids)

    # 차수별(attempt_no) 집계
    weekly_data = _build_attempt_stats(all_logs)

    return {
        "customer": customer,
        "buyers":   [{"company": b["company"], "country": b["country"], "status": b["status"]} for b in buyers],
        "timeline": timeline,
        "progress": progress,
        "stats": {
            "total_buyers": len(buyers),
            "contacted": total_contacted,
            "replied":   total_replied,
            "meetings":  len([m for m in meetings if m["status"] == "scheduled"]),
        },
        "weekly_data": weekly_data,
        "files":     files,
        "meetings":  meetings,
        "inquiries": inquiries,
        "target_countries": target_countries,
    }


def _build_attempt_stats(logs: list) -> list:
    """attempt_no(차수) 기준 누적 컨택/회신 집계
    
    contact_date가 기간(2026.02.09 - 2026.02.23)으로 입력되므로
    주별 대신 차수별(1차/2차/3차...)로 집계
    """
    from collections import defaultdict

    if not logs:
        return []

    # attempt_no별로 바이어 집계
    # contacted: 해당 차수에 컨택된 고유 바이어
    # replied:   해당 차수에서 replied==True인 고유 바이어
    attempt_contacted = defaultdict(set)  # attempt_no -> set of buyer_ids
    attempt_replied   = defaultdict(set)
    attempt_date      = {}  # attempt_no -> contact_date (라벨용)

    for log in logs:
        attempt = log.get("attempt_no")
        if not attempt:
            continue
        buyer_id = log.get("buyer_id")
        attempt_contacted[attempt].add(buyer_id)
        if log.get("replied") is True:
            attempt_replied[attempt].add(buyer_id)
        # 날짜 라벨 저장 (처음 만난 것 사용)
        if attempt not in attempt_date and log.get("contact_date"):
            attempt_date[attempt] = log["contact_date"]

    if not attempt_contacted:
        return []

    # 차수 오름차순 정렬, 누적 합산
    sorted_attempts = sorted(attempt_contacted.keys())
    result = []
    cum_contacted = 0
    cum_replied   = 0

    for attempt in sorted_attempts:
        cum_contacted += len(attempt_contacted[attempt])
        cum_replied   += len(attempt_replied[attempt])

        # 라벨: "1차 (02.09~02.23)" 형식
        date_label = ""
        raw_date = attempt_date.get(attempt, "")
        if raw_date:
            # 2026.02.09 - 2026.02.23  또는  2026.02.09-2026.02.23 형식에서
            # 월.일 ~ 월.일 만 추출
            import re
            parts = re.split(r'\s*[-~]\s*', raw_date.strip())
            def fmt(d: str) -> str:
                nums = re.findall(r'\d+', d)
                if len(nums) >= 3:
                    return f"{int(nums[1]):02d}.{int(nums[2]):02d}"
                return d
            if len(parts) >= 2:
                date_label = f" ({fmt(parts[0])}~{fmt(parts[-1])})"
            elif len(parts) == 1:
                date_label = f" ({fmt(parts[0])})"

        result.append({
            "week":      f"{attempt}차{date_label}",
            "contacted": cum_contacted,
            "replied":   cum_replied,
            "meetings":  0,
        })

    return result


# ── 컨택 로그 엑셀 다운로드 ──────────────────────────────────

@router.get("/contact-logs/{customer_id}/export")
async def export_contact_logs(customer_id: str):
    """고객사의 컨택 로그를 엑셀로 다운로드"""
    from fastapi.responses import StreamingResponse
    import io

    sb = get_supabase()

    # 해당 고객사 바이어 조회
    buyers = sb.table("buyers").select("id,company,country,contact_name,email").eq(
        "customer_id", customer_id
    ).execute().data or []

    buyer_map = {b["id"]: b for b in buyers}
    buyer_ids = list(buyer_map.keys())

    # 컨택 로그 조회 (200개씩 청크로 나눠 조회 — Supabase in_() URL 제한 대응)
    logs = []
    if buyer_ids:
        chunk_size = 200
        for i in range(0, len(buyer_ids), chunk_size):
            chunk = buyer_ids[i:i + chunk_size]
            chunk_logs = sb.table("contact_logs").select(
                "buyer_id, attempt_no, contact_date, replied, reply_content, status, notes"
            ).in_("buyer_id", chunk).order("contact_date").execute().data or []
            logs.extend(chunk_logs)

    # 엑셀 생성
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact Logs"

    # 헤더
    headers = ["No", "Company", "Country", "Contact Name", "Email",
               "Attempt", "Contact Date", "Replied", "Reply Content", "Status", "Notes"]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 컬럼 너비
    col_widths = [6, 28, 15, 20, 30, 10, 15, 10, 40, 15, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 22

    # 데이터
    for i, log in enumerate(logs, 2):
        buyer = buyer_map.get(log.get("buyer_id"), {})
        row_fill = PatternFill("solid", fgColor="F8F7FF") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        row_data = [
            i - 1,
            buyer.get("company", ""),
            buyer.get("country", ""),
            buyer.get("contact_name", ""),
            buyer.get("email", ""),
            log.get("attempt_no", ""),
            log.get("contact_date", ""),
            "Y" if log.get("replied") else "N",
            log.get("reply_content", ""),
            log.get("status", ""),
            log.get("notes", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 9))

    # 스트림으로 반환
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    customer = sb.table("customers").select("name").eq("id", customer_id).single().execute().data
    cname = customer.get("name", "client").replace(" ", "_") if customer else "client"
    filename = f"{cname}_contact_logs_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/dashboard/slug/{slug}")
async def dashboard_by_slug(slug: str):
    """slug 기반 — 반드시 /dashboard/{customer_id} 보다 먼저 선언해야 함"""
    sb = get_supabase()
    result = sb.table("customers").select("id").eq("slug", slug).execute()
    if not result.data or len(result.data) == 0:
        raise HTTPException(status_code=404, detail=f"'{slug}' not found")
    return await _build_dashboard(result.data[0]["id"])


@router.get("/dashboard/{customer_id}")
async def customer_dashboard(customer_id: str):
    return await _build_dashboard(customer_id)


# ── 고객사 생성/조회 ──────────────────────────────────────────

class CustomerCreate(BaseModel):
    name: str
    slug: str
    access_id: Optional[str] = None
    access_password: Optional[str] = None


@router.post("/customers")
async def create_customer(payload: CustomerCreate):
    sb = get_supabase()
    data = {"name": payload.name, "slug": payload.slug}
    if payload.access_id:
        data["access_id"] = payload.access_id
    if payload.access_password:
        # 비번 해시 저장
        data["access_password"] = hashlib.sha256(payload.access_password.encode()).hexdigest()
    result = sb.table("customers").insert(data).execute()
    customer_id = result.data[0]["id"]
    sb.rpc("init_timeline", {"p_customer_id": customer_id}).execute()
    return result.data[0]


@router.get("/customers")
async def list_customers():
    sb = get_supabase()
    # 비번은 노출 안 함
    return sb.table("customers").select("id,name,slug,access_id,created_at").order("name").execute().data


class CustomerAccessUpdate(BaseModel):
    access_id: str
    access_password: str


@router.patch("/customers/{customer_id}/access")
async def update_customer_access(customer_id: str, payload: CustomerAccessUpdate):
    """고객사 아이디/비번 설정 또는 변경"""
    sb = get_supabase()
    hashed = hashlib.sha256(payload.access_password.encode()).hexdigest()
    result = sb.table("customers").update({
        "access_id": payload.access_id,
        "access_password": hashed,
    }).eq("id", customer_id).execute()
    return {"success": True, "access_id": payload.access_id}


# ── 프로필 / 권한 ─────────────────────────────────────────────

@router.get("/profiles")
async def list_profiles():
    sb = get_supabase()
    return sb.table("profiles").select("*, customers(name)").execute().data


class ProfileUpdate(BaseModel):
    role: str   # editor / manager / viewer / customer
    customer_id: Optional[str] = None
    full_name: Optional[str] = None


@router.patch("/profiles/{user_id}")
async def update_profile(user_id: str, payload: ProfileUpdate):
    sb = get_supabase()
    result = sb.table("profiles").update(payload.dict(exclude_none=True)).eq(
        "id", user_id
    ).execute()
    return result.data[0]


@router.delete("/profiles/{user_id}")
async def delete_profile(user_id: str):
    """팀원 삭제 (profiles 테이블에서 제거)"""
    sb = get_supabase()
    sb.table("profiles").delete().eq("id", user_id).execute()
    return {"deleted": True}


# ── 팀원 초대 ─────────────────────────────────────────────────

class InvitePayload(BaseModel):
    email: str
    role: str = "viewer"
    customer_id: Optional[str] = None
    full_name: Optional[str] = None


@router.post("/invite")
async def invite_user(payload: InvitePayload):
    import httpx
    supabase_url = os.getenv("SUPABASE_URL", "")
    service_key  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
    redirect_to  = "https://deta.ai.kr/set-password"
    try:
        resp = httpx.post(
            f"{supabase_url}/auth/v1/invite",
            headers={
                "apikey": service_key,
                "Authorization": f"Bearer {service_key}",
                "Content-Type": "application/json",
            },
            json={
                "email": payload.email,
                "redirect_to": redirect_to,
            },
            timeout=15,
        )
        data = resp.json()
        if resp.status_code == 200 and data.get("id"):
            user_id = data["id"]
            sb = get_supabase()
            sb.table("profiles").upsert({
                "id": user_id,
                "email": payload.email,
                "role": payload.role,
                "customer_id": payload.customer_id or None,
                "full_name": payload.full_name or None,
            }).execute()
            return {"success": True, "message": f"Invitation sent to {payload.email}"}
        else:
            return {"success": False, "error": data.get("msg", "Invite failed")}
    except Exception as e:
        return {"success": False, "error": str(e)}
